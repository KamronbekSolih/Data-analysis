
from openpyxl import load_workbook

  
wb = load_workbook(filename="2. Tug'ilganlar soni-Jami.xlsx")

erkaklar = wb['Erkaklar']
ayollar = wb['Ayollar']

women = {}
men = {}

for col in erkaklar.iter_cols(min_row=4,max_row=5,min_col=5,max_col=27, values_only=True):
    men[col[0]] = col[1]

for col in ayollar.iter_cols(min_row=4,max_row=5,min_col=5,max_col=27, values_only=True):
    women[col[0]] = col[1]

